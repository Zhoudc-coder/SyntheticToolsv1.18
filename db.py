import sqlite3
import datetime
import os
import csv
import json
from pathlib import Path
from openpyxl import load_workbook, Workbook
from config import DB_PATH, RETENTION_DAYS, ensure_log_dir

class LogDatabaseError(Exception):
    """自定义数据库异常"""
    pass


class LogDatabase:
    def __init__(self, db_path=DB_PATH):
        self.db_path = str(db_path)
        self.conn = None

    def connect(self):
        """建立连接并设置SQLite参数，支持并发和网络共享盘"""
        try:
            ensure_log_dir()
        except RuntimeError as e:
            raise LogDatabaseError(str(e))

        try:
            self.conn = sqlite3.connect(self.db_path, timeout=10)
            self.conn.execute("PRAGMA busy_timeout = 10000")
            try:
                self.conn.execute("PRAGMA journal_mode=WAL")
            except sqlite3.OperationalError:
                self.conn.execute("PRAGMA journal_mode=DELETE")
            self.conn.execute("PRAGMA synchronous=NORMAL")
            self.conn.execute("PRAGMA foreign_keys=ON")
        except sqlite3.Error as e:
            raise LogDatabaseError(f"无法连接日志数据库 {self.db_path}：{e}")

    def close(self):
        if self.conn:
            self.conn.close()
            self.conn = None

    def initialize(self):
        """创建新表结构（不执行旧表迁移）"""
        if not self.conn:
            self.connect()
        try:
            self.conn.executescript("""
                CREATE TABLE IF NOT EXISTS merged_boxes (
                    box_code TEXT PRIMARY KEY,
                    original_filename TEXT NOT NULL,
                    shipment_date TEXT NOT NULL,
                    merge_date TEXT NOT NULL,
                    merge_time TEXT NOT NULL,
                    output_filename TEXT NOT NULL,
                    valid_rows INTEGER NOT NULL,
                    total_rows INTEGER NOT NULL,
                    sequence_no INTEGER
                );
                CREATE TABLE IF NOT EXISTS merged_sns (
                    sn TEXT PRIMARY KEY,
                    box_code TEXT NOT NULL,
                    merge_date TEXT NOT NULL,
                    merge_time TEXT NOT NULL,
                    FOREIGN KEY(box_code) REFERENCES merged_boxes(box_code) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_merged_sns_box_code ON merged_sns(box_code);
                CREATE INDEX IF NOT EXISTS idx_merged_sns_merge_date ON merged_sns(merge_date);
                CREATE INDEX IF NOT EXISTS idx_merged_boxes_merge_date ON merged_boxes(merge_date);
                CREATE TABLE IF NOT EXISTS merge_batches (
                    batch_id TEXT PRIMARY KEY,
                    merge_timestamp TEXT NOT NULL,
                    box_codes TEXT NOT NULL,
                    box_count INTEGER NOT NULL DEFAULT 0
                );
            """)
            self.conn.commit()
        except sqlite3.Error as e:
            raise LogDatabaseError(f"初始化数据库表失败：{e}")

    def check_schema_status(self) -> bool:
        """
        检查数据库是否需要迁移（旧表结构存在）。
        返回 True 表示需要迁移，False 表示已是最新结构。
        """
        if not self.conn:
            self.connect()
        # 检查 merged_boxes 表是否旧版
        cursor = self.conn.execute("PRAGMA table_info(merged_boxes)")
        columns = {row[1] for row in cursor.fetchall()}
        if 'package_date' in columns and 'row_count' in columns and 'valid_rows' not in columns:
            return True
        # 检查 merge_batches 表是否存在且缺少 box_count 字段
        cursor = self.conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='merge_batches'")
        if cursor.fetchone():
            cursor = self.conn.execute("PRAGMA table_info(merge_batches)")
            batch_columns = {row[1] for row in cursor.fetchall()}
            if 'box_count' not in batch_columns:
                return True
        return False

    def migrate_schema(self) -> tuple[bool, str]:
        """
        检测并执行旧表迁移。若不需要迁移则返回 (True, "数据库已是最新结构")。
        返回 (是否成功, 消息)。
        """
        if not self.conn:
            self.connect()
        try:
            needs_migration = self.check_schema_status()
            if not needs_migration:
                return True, "数据库已是最新结构，无需更新"

            # 关闭外键约束，以便删除被引用的表
            self.conn.execute("PRAGMA foreign_keys=OFF")

            # 1. 迁移 merged_boxes（如果旧版）
            cursor = self.conn.execute("PRAGMA table_info(merged_boxes)")
            columns = {row[1] for row in cursor.fetchall()}
            if 'package_date' in columns and 'row_count' in columns and 'valid_rows' not in columns:
                self.conn.execute("""
                    CREATE TABLE merged_boxes_new (
                        box_code TEXT PRIMARY KEY,
                        original_filename TEXT NOT NULL,
                        shipment_date TEXT NOT NULL,
                        merge_date TEXT NOT NULL,
                        merge_time TEXT NOT NULL DEFAULT '',
                        output_filename TEXT NOT NULL,
                        valid_rows INTEGER NOT NULL,
                        total_rows INTEGER NOT NULL,
                        sequence_no INTEGER
                    )
                """)
                self.conn.execute("""
                    INSERT INTO merged_boxes_new 
                        (box_code, original_filename, shipment_date, merge_date, merge_time, output_filename, valid_rows, total_rows, sequence_no)
                    SELECT box_code, original_filename, shipment_date, merge_date, '', output_filename, row_count, row_count, sequence_no
                    FROM merged_boxes
                """)
                self.conn.execute("DROP TABLE merged_boxes")
                self.conn.execute("ALTER TABLE merged_boxes_new RENAME TO merged_boxes")
                self.conn.execute("CREATE INDEX IF NOT EXISTS idx_merged_boxes_merge_date ON merged_boxes(merge_date)")

            # 2. 处理 merged_sns：添加 merge_time 列（若不存在）
            cursor = self.conn.execute("PRAGMA table_info(merged_sns)")
            sns_columns = {row[1] for row in cursor.fetchall()}
            if 'merge_time' not in sns_columns:
                self.conn.execute("ALTER TABLE merged_sns ADD COLUMN merge_time TEXT NOT NULL DEFAULT ''")

            # 3. 确保 merge_batches 表存在且包含 box_count 字段
            self.conn.execute("""
                CREATE TABLE IF NOT EXISTS merge_batches (
                    batch_id TEXT PRIMARY KEY,
                    merge_timestamp TEXT NOT NULL,
                    box_codes TEXT NOT NULL,
                    box_count INTEGER NOT NULL DEFAULT 0
                )
            """)
            cursor = self.conn.execute("PRAGMA table_info(merge_batches)")
            batch_columns = {row[1] for row in cursor.fetchall()}
            if 'box_count' not in batch_columns:
                self.conn.execute("ALTER TABLE merge_batches ADD COLUMN box_count INTEGER NOT NULL DEFAULT 0")

            # 重新启用外键
            self.conn.execute("PRAGMA foreign_keys=ON")
            self.conn.commit()
            return True, "数据库结构更新成功"
        except Exception as e:
            self.conn.rollback()
            try:
                self.conn.execute("PRAGMA foreign_keys=ON")
            except:
                pass
            return False, f"数据库迁移失败：{e}"

    def cleanup_old_logs(self, days=RETENTION_DAYS):
        if not self.conn:
            self.connect()
        cutoff = (datetime.date.today() - datetime.timedelta(days=days)).isoformat()
        self.conn.execute("DELETE FROM merged_sns WHERE merge_date < ?", (cutoff,))
        self.conn.execute("DELETE FROM merged_boxes WHERE merge_date < ?", (cutoff,))
        self.conn.commit()

    def load_existing_box_codes(self):
        if not self.conn:
            self.connect()
        cursor = self.conn.execute("SELECT box_code FROM merged_boxes")
        return {row[0] for row in cursor.fetchall()}

    def load_existing_sns(self):
        if not self.conn:
            self.connect()
        cursor = self.conn.execute("SELECT sn FROM merged_sns")
        return {row[0] for row in cursor.fetchall()}

    def insert_merge_records(self, box_records, sn_records, batch_id: str):
        """
        插入合并记录，并同时记录批次信息（box_codes 列表和 box_count）。
        box_records 和 sn_records 均不包含 batch_id 字段。
        """
        if not self.conn:
            self.connect()
        try:
            self.conn.execute("BEGIN IMMEDIATE")
            self.conn.executemany(
                "INSERT INTO merged_boxes "
                "(box_code, original_filename, shipment_date, merge_date, merge_time, "
                " output_filename, valid_rows, total_rows, sequence_no) "
                "VALUES (?,?,?,?,?,?,?,?,?)",
                box_records
            )
            self.conn.executemany(
                "INSERT INTO merged_sns (sn, box_code, merge_date, merge_time) VALUES (?,?,?,?)",
                sn_records
            )
            # 记录批次信息
            box_codes_list = [rec[0] for rec in box_records]
            box_count = len(box_codes_list)
            merge_timestamp = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            self.conn.execute(
                "INSERT INTO merge_batches (batch_id, merge_timestamp, box_codes, box_count) VALUES (?,?,?,?)",
                (batch_id, merge_timestamp, json.dumps(box_codes_list), box_count)
            )
            self.conn.commit()
            return True
        except sqlite3.IntegrityError:
            self.conn.rollback()
            return False
        except Exception:
            self.conn.rollback()
            raise

    def delete_merge_records(self, box_codes):
        if not self.conn:
            self.connect()
        try:
            self.conn.execute("BEGIN IMMEDIATE")
            self.conn.executemany(
                "DELETE FROM merged_boxes WHERE box_code = ?",
                [(bc,) for bc in box_codes]
            )
            self.conn.commit()
        except Exception:
            self.conn.rollback()
            raise

    def get_latest_batch(self):
        """获取最近一次合并的批次信息（batch_id 和 box_codes 列表），不依赖 box_count 字段"""
        if not self.conn:
            self.connect()
        cursor = self.conn.execute(
            "SELECT batch_id, box_codes FROM merge_batches ORDER BY merge_timestamp DESC LIMIT 1"
        )
        row = cursor.fetchone()
        if not row:
            return None
        return {
            'batch_id': row[0],
            'box_codes': json.loads(row[1])
        }

    def undo_last_merge(self) -> int:
        """
        撤销最近一次合并操作，删除该批次插入的所有箱码记录（自动级联删除SN）。
        同时删除批次记录本身。
        返回删除的箱码记录数。
        """
        if not self.conn:
            self.connect()
        batch_info = self.get_latest_batch()
        if not batch_info:
            return 0
        batch_id = batch_info['batch_id']
        box_codes = batch_info['box_codes']
        try:
            self.conn.execute("BEGIN IMMEDIATE")
            self.conn.executemany(
                "DELETE FROM merged_boxes WHERE box_code = ?",
                [(bc,) for bc in box_codes]
            )
            self.conn.execute("DELETE FROM merge_batches WHERE batch_id = ?", (batch_id,))
            self.conn.commit()
            return len(box_codes)
        except Exception:
            self.conn.rollback()
            raise

    def get_batches(self, date_filter: str = None):
        """
        查询批次记录，支持按日期过滤（格式 YYYY-MM-DD）。
        返回列表，每个元素包含：batch_id, merge_timestamp, box_count, box_codes_preview。
        box_codes_preview 为前3个箱码，用于在界面显示概要。
        """
        if not self.conn:
            self.connect()
        if date_filter:
            query = """
                SELECT batch_id, merge_timestamp, box_count, box_codes
                FROM merge_batches
                WHERE date(merge_timestamp) = ?
                ORDER BY merge_timestamp DESC
            """
            params = (date_filter,)
        else:
            query = """
                SELECT batch_id, merge_timestamp, box_count, box_codes
                FROM merge_batches
                ORDER BY merge_timestamp DESC
            """
            params = ()
        cursor = self.conn.execute(query, params)
        results = []
        for row in cursor:
            batch_id, merge_timestamp, box_count, box_codes_json = row
            # 解析 JSON 并生成预览
            try:
                box_codes = json.loads(box_codes_json)
                if len(box_codes) > 3:
                    preview = ', '.join(box_codes[:3]) + '...'
                else:
                    preview = ', '.join(box_codes)
            except:
                preview = box_codes_json[:50]  # 若解析失败，截取前50字符
            results.append({
                'batch_id': batch_id,
                'merge_timestamp': merge_timestamp,
                'box_count': box_count,
                'box_codes_preview': preview
            })
        return results

    def undo_batch_by_id(self, batch_id: str) -> int:
        """
        根据批次号撤销该批次的所有记录（箱码及关联SN），并删除批次记录本身。
        返回删除的箱码数，若批次不存在或没有箱码，返回0。
        """
        if not self.conn:
            self.connect()
        # 获取该批次的箱码列表
        cursor = self.conn.execute(
            "SELECT box_codes FROM merge_batches WHERE batch_id = ?", (batch_id,)
        )
        row = cursor.fetchone()
        if not row:
            return 0
        box_codes = json.loads(row[0])
        if not box_codes:
            return 0
        try:
            self.conn.execute("BEGIN IMMEDIATE")
            # 删除箱码记录（级联删除SN）
            self.conn.executemany(
                "DELETE FROM merged_boxes WHERE box_code = ?",
                [(bc,) for bc in box_codes]
            )
            # 删除批次记录
            self.conn.execute("DELETE FROM merge_batches WHERE batch_id = ?", (batch_id,))
            self.conn.commit()
            return len(box_codes)
        except Exception:
            self.conn.rollback()
            raise

    def export_boxes_to_excel(self, output_path, date_filter: str = None) -> int:
        if not self.conn:
            self.connect()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("BoxRecords")
        ws.append([
            "box_code", "original_filename", "shipment_date",
            "merge_date", "merge_time", "output_filename",
            "valid_rows", "total_rows", "sequence_no"
        ])
        count = 0
        if date_filter:
            query = (
                "SELECT box_code, original_filename, shipment_date, "
                "merge_date, merge_time, output_filename, valid_rows, total_rows, sequence_no "
                "FROM merged_boxes WHERE shipment_date = ? ORDER BY sequence_no"
            )
            params = (date_filter,)
        else:
            query = (
                "SELECT box_code, original_filename, shipment_date, "
                "merge_date, merge_time, output_filename, valid_rows, total_rows, sequence_no "
                "FROM merged_boxes ORDER BY shipment_date DESC, sequence_no"
            )
            params = ()
        cursor = self.conn.execute(query, params)
        for row in cursor:
            ws.append(list(row))
            count += 1
        wb.save(output_path)
        return count

    def export_sns_to_excel(self, output_path, date_filter: str = None) -> int:
        if not self.conn:
            self.connect()
        wb = Workbook(write_only=True)
        ws = wb.create_sheet("SNRecords")
        ws.append(["SN", "box_code", "merge_date", "merge_time"])
        count = 0
        if date_filter:
            query = """
                    SELECT s.sn, s.box_code, s.merge_date, s.merge_time
                    FROM merged_sns s
                    INNER JOIN merged_boxes b ON s.box_code = b.box_code
                    WHERE b.shipment_date = ? ORDER BY s.sn
                    """
            params = (date_filter,)
        else:
            query = """
                    SELECT s.sn, s.box_code, s.merge_date, s.merge_time
                    FROM merged_sns s ORDER BY s.merge_date DESC, s.sn
                    """
            params = ()
        cursor = self.conn.execute(query, params)
        for row in cursor:
            ws.append(list(row))
            count += 1
        wb.save(output_path)
        return count

    def export_sns_with_data(self, output_path: str, data_file_path: str, date_filter: str = None) -> int:
        if not self.conn:
            self.connect()

        # 从数据库加载 SN 映射
        sn_map = {}
        if date_filter:
            query = """
                    SELECT s.sn, s.box_code, s.merge_date, s.merge_time
                    FROM merged_sns s
                    INNER JOIN merged_boxes b ON s.box_code = b.box_code
                    WHERE b.shipment_date = ?
                    """
            params = (date_filter,)
        else:
            query = "SELECT sn, box_code, merge_date, merge_time FROM merged_sns"
            params = ()
        cursor = self.conn.execute(query, params)
        for sn, box_code, merge_date, merge_time in cursor:
            sn_map[sn] = (box_code, merge_date, merge_time)

        data_path = Path(data_file_path)
        suffix = data_path.suffix.lower()
        if suffix == '.xlsx':
            wb_in = load_workbook(filename=str(data_path), read_only=True, data_only=True)
            ws_in = wb_in.active
            try:
                title_row = next(ws_in.iter_rows(min_row=1, max_row=1, values_only=True))
            except StopIteration:
                wb_in.close()
                raise ValueError("数据文件没有标题行")
            title_len = len(title_row)
            data_iter = ws_in.iter_rows(min_row=2, values_only=True)
            close_func = wb_in.close
        elif suffix == '.csv':
            encodings = ['utf-8-sig', 'utf-8', 'gbk']
            file_handle = None
            reader = None
            for enc in encodings:
                try:
                    file_handle = open(data_path, 'r', encoding=enc, newline='')
                    reader = csv.reader(file_handle)
                    try:
                        title_row = next(reader)
                    except StopIteration:
                        file_handle.close()
                        raise ValueError("CSV 文件为空")
                    title_len = len(title_row)
                    data_iter = reader
                    close_func = file_handle.close
                    break
                except UnicodeDecodeError:
                    if file_handle:
                        file_handle.close()
                    continue
            if reader is None:
                raise ValueError("无法解码 CSV 文件，请使用 UTF-8 或 GBK 编码")
        else:
            raise ValueError(f"不支持的文件类型: {suffix}")

        wb_out = Workbook(write_only=True)
        ws_out = wb_out.create_sheet("SNData")
        new_title = list(title_row) + ["box_code", "merge_date", "merge_time"]
        ws_out.append(new_title)

        exported_count = 0
        for row in data_iter:
            if not row:
                continue
            sn_raw = row[0] if len(row) > 0 else None
            sn = str(sn_raw).strip() if sn_raw is not None else ''
            if sn in sn_map:
                row_list = list(row)
                if len(row_list) < title_len:
                    row_list.extend([None] * (title_len - len(row_list)))
                elif len(row_list) > title_len:
                    row_list = row_list[:title_len]
                box_code, merge_date, merge_time = sn_map[sn]
                row_list.extend([box_code, merge_date, merge_time])
                ws_out.append(row_list)
                exported_count += 1

        close_func()
        wb_out.save(output_path)
        return exported_count

    def get_database_info(self):
        if not self.conn:
            self.connect()
        boxes_count = self.conn.execute("SELECT COUNT(*) FROM merged_boxes").fetchone()[0]
        sns_count = self.conn.execute("SELECT COUNT(*) FROM merged_sns").fetchone()[0]
        mtime = os.path.getmtime(self.db_path)
        mtime_str = datetime.datetime.fromtimestamp(mtime).strftime('%Y-%m-%d %H:%M:%S')
        return boxes_count, sns_count, mtime_str

    def get_latest_boxes(self, limit=10):
        if not self.conn:
            self.connect()
        cursor = self.conn.execute(
            "SELECT box_code, original_filename, shipment_date, "
            "merge_date, merge_time, output_filename, valid_rows, total_rows, sequence_no "
            "FROM merged_boxes ORDER BY merge_date DESC, rowid DESC LIMIT ?",
            (limit,)
        )
        columns = ["box_code", "original_filename", "shipment_date",
                   "merge_date", "merge_time", "output_filename", "valid_rows", "total_rows", "sequence_no"]
        return [dict(zip(columns, row)) for row in cursor.fetchall()]