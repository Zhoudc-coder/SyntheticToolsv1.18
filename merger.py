import csv
import datetime
import uuid
from pathlib import Path
from openpyxl import load_workbook, Workbook

from config import DB_PATH, RETENTION_DAYS
from db import LogDatabase
from file_parser import validate_sn


def _read_rows(file_path: Path):
    """
    根据文件后缀读取标题行和数据行迭代器。
    支持 .xlsx 和 .csv 格式。
    返回 (title_row, data_iter)，其中 data_iter 是一个生成器，逐行产出数据（不包含标题）。
    注意：调用方应在使用完 data_iter 后调用 close() 方法释放资源。
    """
    suffix = file_path.suffix.lower()
    if suffix == '.xlsx':
        wb = load_workbook(filename=str(file_path), read_only=True, data_only=True)
        ws = wb.active
        try:
            title_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
        except StopIteration:
            wb.close()
            raise ValueError("文件没有标题行")

        def data_iter():
            try:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    yield row
            finally:
                wb.close()

        return title_row, data_iter()

    elif suffix == '.csv':
        # 尝试不同编码
        encodings = ['utf-8-sig', 'utf-8', 'gbk']
        for enc in encodings:
            try:
                f = open(file_path, 'r', encoding=enc, newline='')
                reader = csv.reader(f)
                try:
                    title_row = next(reader)
                except StopIteration:
                    f.close()
                    raise ValueError("CSV 文件为空")

                def data_iter():
                    try:
                        for row in reader:
                            yield row
                    finally:
                        f.close()

                return title_row, data_iter()
            except UnicodeDecodeError:
                continue
        raise ValueError("无法解码 CSV 文件，请使用 UTF-8 或 GBK 编码")
    else:
        raise ValueError(f"不支持的文件类型: {suffix}")


def _get_unique_filename(directory: Path, base_name: str, extension: str = '.xlsx') -> Path:
    """生成不重复的文件名，自动添加 (1), (2) 等后缀"""
    final_path = directory / (base_name + extension)
    counter = 1
    while final_path.exists():
        final_path = directory / (f"{base_name} ({counter})" + extension)
        counter += 1
    return final_path


def perform_merge(file_infos, output_dir: Path, output_filename_prefix: str,
                  shipment_date: str, log_func=print,
                  progress_callback=None,
                  auto_export_boxes: bool = True) -> tuple[bool, list[str], str]:
    """
    执行合并操作，按版本分组生成多个汇总文件。
    文件名使用用户选择的出库日期（而非合并当天日期）。
    自动添加后缀避免覆盖。
    当 auto_export_boxes=True 时，同时为每个版本生成箱码记录文件。
    进度通过 progress_callback 回调，参数为字符串消息。
    """
    if not file_infos:
        return False, [], "没有待合并的文件"

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    # 生成本次合并的批次ID（20位）
    batch_id = (datetime.datetime.now().strftime('%Y%m%d%H%M%S%f') + uuid.uuid4().hex)[:20]

    # 初始化日志数据库
    db = LogDatabase(DB_PATH)
    try:
        db.initialize()
        db.cleanup_old_logs(RETENTION_DAYS)
        # 加载历史箱码和文件名映射
        existing_box_codes = set()
        existing_box_filename = {}
        cursor = db.conn.execute("SELECT box_code, original_filename FROM merged_boxes")
        for box_code, original_filename in cursor:
            existing_box_codes.add(box_code)
            existing_box_filename[box_code] = original_filename

        # 加载历史SN及SN到箱码的映射
        existing_sns = set()
        sn_to_box_code = {}
        cursor = db.conn.execute("SELECT sn, box_code FROM merged_sns")
        for sn, box_code in cursor:
            existing_sns.add(sn)
            sn_to_box_code[sn] = box_code

    except Exception as e:
        return False, [], f"日志数据库初始化失败：{e}"

    # 实际合并日期（仅用于日志记录）
    merge_date = datetime.date.today()
    merge_date_iso = merge_date.isoformat()

    # 使用用户选择的出库日期生成文件名日期部分
    try:
        shipment_date_obj = datetime.date.fromisoformat(shipment_date)
    except ValueError:
        return False, [], f"出库日期格式错误：{shipment_date}"
    shipment_date_str = shipment_date_obj.strftime('%y%m%d')

    # 按版本分组
    version_groups = {}
    for file_path, box_info in file_infos:
        version = box_info.version
        version_groups.setdefault(version, []).append((file_path, box_info))

    total_files = len(file_infos)
    processed_files = 0
    total_versions = len(version_groups)
    version_index = 0

    # 记录日志所需的数据
    box_records = []          # 全部箱码记录（用于数据库）
    sn_records = []           # 全部 SN 记录
    seen_sns = set()          # 全局 SN 去重（本次合并）
    seen_sn_source = {}       # 记录本次合并中每个SN首次出现的文件名
    generated_files = []      # 生成的汇总文件名列表
    version_stats = {}        # 版本统计信息

    # 循环处理每个版本
    for version, files in version_groups.items():
        version_index += 1
        log_func(f"开始处理版本 v{version}，共 {len(files)} 个文件")
        version_stats[version] = {'files': 0, 'total_rows': 0, 'valid_rows': 0}

        # 当前版本的箱码记录（包含所有成功处理的文件，有效行数可能为0）
        version_box_records = []

        # 生成唯一的输出文件名（自动添加后缀避免覆盖）
        base_filename = f"{output_filename_prefix}-{shipment_date_str}-v{version}"
        final_path = _get_unique_filename(output_dir, base_filename, '.xlsx')
        final_filename = final_path.name

        temp_path = output_dir / (final_filename + '.tmp')

        # 读取该版本第一个文件的标题
        try:
            first_file = files[0][0]
            title_row, _ = _read_rows(first_file)
            title_len = len(title_row)
        except Exception as e:
            log_func(f"读取版本 v{version} 第一个文件标题失败：{first_file.name} - {e}")
            continue

        # 创建该版本的汇总工作簿
        out_wb = Workbook(write_only=True)
        out_ws = out_wb.create_sheet(title="汇总")
        out_ws.append(list(title_row))

        for file_path, box_info in files:
            processed_files += 1
            if progress_callback:
                progress_callback(f"正在合并数据：{processed_files}/{total_files}")

            log_func(f"处理文件：{file_path.name} (v{version})")
            data_iter = None
            try:
                file_title, data_iter = _read_rows(file_path)

                # 标题不一致仅警告，不跳过
                if len(file_title) != title_len:
                    log_func(
                        f"警告：文件标题列数不一致（{len(file_title)} 列 vs {title_len} 列），"
                        f"将按基准列数 {title_len} 对齐，继续处理：{file_path.name}"
                    )
                elif any(a != b for a, b in zip(file_title, title_row)):
                    log_func(
                        f"警告：文件标题内容存在差异，可能影响列对应关系，"
                        f"但将继续处理：{file_path.name}"
                    )

                file_total_rows = 0
                row_count = 0
                merge_time_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M')

                for row in data_iter:
                    if not row:
                        continue
                    file_total_rows += 1

                    sn_raw = row[0] if len(row) > 0 else None
                    sn = str(sn_raw).strip() if sn_raw is not None else ''
                    if not validate_sn(sn):
                        log_func(f"警告：SN格式不符合，跳过行：{sn or '<空>'}")
                        continue

                    # 检查重复
                    if sn in existing_sns or sn in seen_sns:
                        if sn in existing_sns:
                            hist_box = sn_to_box_code.get(sn, '未知箱码')
                            hist_file = existing_box_filename.get(hist_box, '未知文件')
                            log_func(
                                f"警告：SN重复（与历史记录重复），跳过：{sn}，"
                                f"历史箱码：{hist_box}，历史文件：{hist_file}"
                            )
                        else:
                            src_file = seen_sn_source.get(sn, '未知文件')
                            log_func(
                                f"警告：SN重复（本次合并已存在），跳过：{sn}，"
                                f"首次出现文件：{src_file}"
                            )
                        continue

                    # 有效行处理
                    row_list = list(row)
                    if len(row_list) < title_len:
                        row_list.extend([None] * (title_len - len(row_list)))
                    elif len(row_list) > title_len:
                        log_func(f"警告：行数据列数超过标题，截断：{sn}")
                        row_list = row_list[:title_len]

                    out_ws.append(row_list)
                    seen_sns.add(sn)
                    seen_sn_source[sn] = file_path.name
                    sn_records.append((sn, box_info.box_code, merge_date_iso, merge_time_str))
                    row_count += 1

                # 记录箱码（无论有效行数是否为0）
                box_record = (
                    box_info.box_code,
                    file_path.name,
                    shipment_date,
                    merge_date_iso,
                    merge_time_str,
                    final_filename,
                    row_count,          # valid_rows
                    file_total_rows,    # total_rows
                    box_info.sequence_no
                )
                box_records.append(box_record)
                version_box_records.append(box_record)

                version_stats[version]['files'] += 1
                version_stats[version]['total_rows'] += file_total_rows
                version_stats[version]['valid_rows'] += row_count

            except Exception as e:
                log_func(f"错误：读取文件失败，跳过：{file_path.name} - {e}")
                continue
            finally:
                if data_iter is not None:
                    try:
                        data_iter.close()
                    except Exception:
                        pass

        # 保存临时汇总文件
        try:
            out_wb.save(str(temp_path))
            log_func(f"版本 v{version} 临时汇总文件已生成：{temp_path}")
        except Exception as e:
            log_func(f"版本 v{version} 写入临时汇总文件失败：{e}")
            continue

        # 替换最终汇总文件
        try:
            temp_path.replace(final_path)
            generated_files.append(final_filename)
            log_func(f"版本 v{version} 汇总完成：{final_filename}，有效 {version_stats[version]['valid_rows']} 行，总 {version_stats[version]['total_rows']} 行")
        except Exception as e:
            log_func(f"版本 v{version} 替换最终文件失败：{e}")
            temp_path.unlink(missing_ok=True)
            continue

        # 自动生成箱码记录文件（包含所有处理的箱码，即使有效行0）
        if auto_export_boxes and version_box_records:
            if progress_callback:
                progress_callback(f"正在自动导出箱码记录：{version_index}/{total_versions}")

            box_record_base = f"{output_filename_prefix}-{shipment_date_str}-v{version}-箱码记录"
            box_record_path = _get_unique_filename(output_dir, box_record_base, '.xlsx')
            try:
                wb_box = Workbook(write_only=True)
                ws_box = wb_box.create_sheet("箱码记录")
                ws_box.append([
                    "box_code", "original_filename", "shipment_date",
                    "merge_date", "merge_time", "output_filename",
                    "valid_rows", "total_rows", "sequence_no"
                ])
                for rec in version_box_records:
                    ws_box.append(list(rec))
                wb_box.save(str(box_record_path))
                log_func(f"箱码记录文件已生成：{box_record_path.name}")
                generated_files.append(box_record_path.name)
            except Exception as e:
                log_func(f"版本 v{version} 生成箱码记录失败：{e}")

    if not generated_files:
        return False, [], "没有生成任何汇总文件（可能所有版本均处理失败）"

    # 写入日志
    try:
        insert_ok = db.insert_merge_records(box_records, sn_records, batch_id)
    except Exception as e:
        for fname in generated_files:
            (output_dir / fname).unlink(missing_ok=True)
        return False, [], f"日志写入异常：{e}"

    if not insert_ok:
        for fname in generated_files:
            (output_dir / fname).unlink(missing_ok=True)
        return False, [], "日志写入失败：箱码或SN与已有记录冲突（可能其他电脑已合并），请重新操作"

    # 构建统计信息
    total_scanned_files = len(file_infos)
    total_valid_files = sum(stats['files'] for stats in version_stats.values())
    total_rows_all = sum(stats['total_rows'] for stats in version_stats.values())
    total_valid_rows = sum(stats['valid_rows'] for stats in version_stats.values())

    summary_lines = [
        f"合并成功！共扫描 {total_scanned_files} 个文件，其中有效文件 {total_valid_files} 个。总共 {total_rows_all} 行数据，其中有效数据 {total_valid_rows} 行。生成 {len(generated_files)} 个文件（含汇总文件和箱码记录）。"
    ]
    for ver in sorted(version_stats.keys()):
        stats = version_stats[ver]
        summary_lines.append(
            f"  版本 v{ver}: 文件数 {stats['files']}，总共 {stats['total_rows']} 行，有效数据 {stats['valid_rows']} 行"
        )

    summary_msg = "\n".join(summary_lines)
    log_func(summary_msg)

    # 输出合并完成时间和分隔符
    finish_time_str = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    log_func("=" * 60)
    log_func(f"合并完成时间：{finish_time_str}")
    log_func("=" * 60)

    return True, generated_files, summary_msg